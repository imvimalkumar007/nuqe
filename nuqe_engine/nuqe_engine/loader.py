"""
M1: Library loader.

Reads Nuqe_Obligation_Library.xlsx from disk and emits a list of
RawObligationRow objects (one per row in the obligation_library sheet,
excluding the header row).

Filtering happens here: only rows with review_status='approved' are loaded
by default, so the engine never operates on draft or peer_review obligations.
The validator (M2) is responsible for parsing the structured sub-fields.

This module is read-only. It does not validate, transform, or persist.
"""

from __future__ import annotations

import logging
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

from nuqe_engine.schema import RawObligationRow, ReviewStatus, column_order

logger = logging.getLogger(__name__)


# Excel's 1900-based serial date system, accounting for the 1900 leap year bug
# that means Excel treats 1900 as a leap year (it isn't). Day 1 in Excel is
# 1900-01-01; we use 1899-12-30 as the epoch so that integer 1 maps to 1900-01-01
# while integer 60 (Excel's spurious Feb 29 1900) maps to 1900-02-28 correctly.
_EXCEL_EPOCH = date(1899, 12, 30)


class LoaderError(Exception):
    """Raised when the spreadsheet cannot be read or its structure is invalid."""


def _excel_serial_to_date(serial: int | float) -> date:
    """Convert an Excel serial date number to a Python date."""
    return _EXCEL_EPOCH + timedelta(days=int(serial))


# Columns where openpyxl may return cell values of various types depending on
# how the cell was written. The loader is robust to ISO strings, datetime,
# date, and Excel serial integers in these columns.
_DATE_COLUMNS = frozenset({"effective_from", "effective_to"})


def _normalise_cell(value: object, column_name: str | None = None) -> object:
    """
    Convert openpyxl cell values to plain Python types we expect downstream.

    Date columns may arrive as:
      - ISO date strings ('2011-09-01') from older programmatic writes
      - datetime objects when the cell has a date number-format
      - date objects (rare)
      - bare integers when the cell stores an Excel serial date without
        a date number-format (this happens when openpyxl is given a date()
        and not told to apply a date style)

    For non-date columns, integers and strings pass through unchanged.
    """
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        # Try ISO date for date columns
        if column_name in _DATE_COLUMNS:
            try:
                return date.fromisoformat(stripped)
            except ValueError:
                # Not a date string. Fall through; pydantic will report.
                return stripped
        return stripped
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if column_name in _DATE_COLUMNS and isinstance(value, int | float):
        # Excel serial date in an unstyled cell
        return _excel_serial_to_date(value)
    return value


def load_library(
    path: str | Path,
    *,
    sheet_name: str = "obligation_library",
    approved_only: bool = True,
) -> list[RawObligationRow]:
    """
    Load all obligations from the given xlsx file.

    Args:
        path: Path to Nuqe_Obligation_Library.xlsx.
        sheet_name: Sheet to read. Defaults to 'obligation_library'.
        approved_only: If True (default), only rows with review_status='approved'
            are returned. This is the engine's normal operating mode: it never
            consumes draft or peer_review obligations.

    Returns:
        List of RawObligationRow, one per data row. Strings are stripped;
        empty strings are normalised to None for nullable columns.

    Raises:
        LoaderError: If the file is missing, the sheet is missing, the header
            row does not match the canonical column order, or any cell has the
            wrong type for its column.
    """
    p = Path(path)
    if not p.exists():
        raise LoaderError(f"Library file not found: {p}")
    if not p.is_file():
        raise LoaderError(f"Library path is not a file: {p}")

    try:
        wb = load_workbook(filename=p, data_only=True, read_only=True)
    except Exception as exc:
        raise LoaderError(f"Failed to open workbook {p}: {exc}") from exc

    if sheet_name not in wb.sheetnames:
        raise LoaderError(
            f"Sheet '{sheet_name}' not found in {p}. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)

    # Header row check
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise LoaderError(f"Sheet '{sheet_name}' is empty (no header row)") from None

    expected = column_order()
    header = [h for h in header_row if h is not None]
    if header[: len(expected)] != expected:
        raise LoaderError(
            "Spreadsheet header does not match the canonical 24-column schema.\n"
            f"Expected: {expected}\n"
            f"Got: {header[: len(expected)]}"
        )

    rows: list[RawObligationRow] = []
    for row_idx, data_row in enumerate(rows_iter, start=2):
        # Skip fully empty rows
        if all(cell is None for cell in data_row):
            continue

        # Map by position (the columns are fixed by the Method)
        normalised = [
            _normalise_cell(c, column_name=col)
            for c, col in zip(data_row[: len(expected)], expected, strict=False)
        ]

        # Filter early on review_status (column 24)
        review_status = normalised[expected.index("review_status")]
        if approved_only and review_status != ReviewStatus.APPROVED.value:
            continue

        # Build the raw row
        kwargs = dict(zip(expected, normalised, strict=False))
        try:
            raw = RawObligationRow(**kwargs)  # type: ignore[arg-type]
        except Exception as exc:
            raise LoaderError(
                f"Row {row_idx} ({kwargs.get('obligation_id', '<no id>')}) "
                f"failed shape validation: {exc}"
            ) from exc

        # Stash provenance for downstream error messages. Pydantic v2 doesn't
        # serialise private attrs by default, which is fine: this is internal.
        object.__setattr__(raw, "_source_row_number", row_idx)
        rows.append(raw)

    wb.close()
    logger.info(
        "Loaded %d obligations from %s (approved_only=%s)",
        len(rows),
        p.name,
        approved_only,
    )
    return rows


def load_all_statuses(path: str | Path) -> dict[str, list[RawObligationRow]]:
    """
    Load every row and group by review_status. Useful for reporting and CLI.

    Returns:
        Dict mapping review_status string to list of RawObligationRow.
    """
    rows_by_status: dict[str, list[RawObligationRow]] = {
        s.value: [] for s in ReviewStatus
    }
    all_rows = load_library(path, approved_only=False)
    for r in all_rows:
        rows_by_status.setdefault(r.review_status, []).append(r)
    return rows_by_status
